#! python3
#
# SNReport2L-B.py -- Read the Service Now Assessment Report
# Two Line Report version with Estimated Completion Date
#
# import OPENPYXL Functions
#
import openpyxl
from openpyxl.styles import NamedStyle, Alignment
#from openpyxl.comments import Comment
#
# Import OS Functions
#
import os
#
# Import TK GUI FUNCTIONS
#
import tkinter
from tkinter import messagebox
from tkinter import filedialog
from tkinter import Label

#
# Starting TK and hiding the main window
#
root = tkinter.Tk()
root.attributes('-fullscreen', True)
#root.withdraw()
LABEL = Label(root, text="Sevice Now Report Processing!")
LABEL.pack()

#
# Starting the first phase of processing
# Get the input file name to process
#
SNFile = filedialog.askopenfile(parent=root,mode='rb',title='Enter the File Name to process')
print('Opening Workbook ', SNFile.name)
Output = 'Opening Workbook'
LABEL = Label(root, text=Output)
LABEL.pack()
Output = SNFile.name
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
#
# Open the input file
#
wb = openpyxl.load_workbook(SNFile)
#
# Open the worksheet
#
#print('Opening Worksheet')
sheet = wb.get_sheet_by_name('Page 1')
#
# Save the temporary file
#
#print('Saving SNReportTemp.xlsx.')
Output = 'Saving SNReportTemp.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb.save('SNReportTemp.xlsx')
#
# Close the temporary file
#
wb.close()
#
# Start Excel with the temporary file
# Instructions for the expected Excel error
#
Output = 'Starting Excel with SNReportTemp.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
messagebox.showinfo("Starting Excel", "Answer YES to the Error, and then close the Error Log Box")
os.startfile('SNReportTemp.xlsx')
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
#
messagebox.showinfo("In Excel", 'Remove columns B, D, F, I.\n\nCut Column F and Insert at D.\n\nInsert One column at G.\n\nInsert One column at K.\n\nStarting at column M, Insert 9 Columns (M thru U).\n\nSave as SNReportTemp.xlsx, Overwriting the file!\n\n\n\nExit Excel!!!')
#
# Start the second phase of processing
# Open the temporary file
#
#print('Opening File SNReportTemp.xlsx')
Output = 'Opening File SNReportTemp.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb = openpyxl.load_workbook('SNReportTemp.xlsx')
#
# Open the worksheet
#
#print('Opening Worksheet')
Output = 'Opening Worksheet'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
sheet = wb.get_sheet_by_name('Page 1')
#
# Adding the names for the title row
#
#print('Adding New Row 1 Names')
Output = 'Ading new Row 1 Names'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
sheet['a1'] = 'Assigned To RT'
sheet['b1'] = 'Engagement RT'
sheet['c1'] = 'Request RT'
sheet['d1'] = 'Task RT'
sheet['e1'] = 'Created RT'
sheet['f1'] = 'Completion RT'
sheet['g1'] = 'Type RT'
sheet['h1'] = 'Description RT'
sheet['i1'] = 'Status RT'
sheet['j1'] = 'PVID RT'
sheet['k1'] = 'Feature RT'
sheet['l1'] = 'Estimate RT'
sheet['m1'] = 'Name FT'
sheet['n1'] = 'Integration FT'
sheet['o1'] = 'Status FT'
sheet['p1'] = 'VP FT'
sheet['q1'] = 'IT Vertical FT'
sheet['r1'] = 'IT Portfolio FT'
sheet['s1'] = 'MD FT'
sheet['t1'] = 'Sponsor Division FT'
sheet['u1'] = 'Sponsor FT'
sheet['v1'] = 'Requested By Name RT'
sheet['w1'] = 'Requested By Email RT'
sheet['x1'] = 'Requested For Name RT'
sheet['y1'] = 'Requested For Email RT'
sheet['z1'] = 'Approve Date FT'
sheet['aa1'] = 'Approve Start Date FT'
sheet['ab1'] = 'Approve End Date FT'
sheet['ac1'] = 'Planned Start Date FT'
sheet['ad1'] = 'Planned End Date FT'
sheet['ae1'] = 'Project Manager FT'
sheet['af1'] = 'Project Analyst FT'
sheet['ag1'] = 'Project Category FT'
#
# Changing the beginning of the Descriotion to the short version
#
# Change Deliver
#print('Changing to Deliver')
Output = 'Changing to Deliver'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(1,sheet.max_row+1):
    Deliver = sheet.cell(row=i, column=8).value
    if 'Deliver network assessment'in Deliver:
        Deliver2 = Deliver.split('-')
        Type = 'Deliver'
        Deliver2[0] = ''
        Deliver3 = (' ').join(Deliver2)
        Deliver4 = Deliver3.lstrip()
        sheet.cell(row=i, column=7).value = Type
        sheet.cell(row=i, column=8).value = Deliver4
# Change Estimate
#print('Changing to Estimate')
Output = 'Changing to Estimate'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(1,sheet.max_row+1):
    Estimate = sheet.cell(row=i, column=8).value
    if 'Build assessment estimate'in Estimate:
        Estimate2 = Estimate.split('-')
        Type = 'Estimate'
        Estimate2[0] = ''
        Estimate3 = (' ').join(Estimate2)
        Estimate4 = Estimate3.lstrip()
        sheet.cell(row=i, column=7).value = Type
        sheet.cell(row=i, column=8).value = Estimate4
# Change Planning
#print('Changing to Planning')
Output = 'Changing to Planning'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(1,sheet.max_row+1):
    Planning = sheet.cell(row=i, column=8).value
    if 'Resource & Demand planning for'in Planning:
        Planning2 = Planning.split('-')
        Type = 'Planning'
        Planning2[0] = ''
        Planning3 = (' ').join(Planning2)
        Planning4 = Planning3.lstrip()
        sheet.cell(row=i, column=7).value = Type
        sheet.cell(row=i, column=8).value = Planning4
Output = 'Changing to Review'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(1,sheet.max_row+1):
    Review = sheet.cell(row=i, column=8).value
    if 'Review'in Review:
        Review2 = Review.split('-')
        Type = 'Review'
        Review2[0] = ''
        Review3 = (' ').join(Review2)
        Review4 = Review.lstrip()
        sheet.cell(row=i, column=7).value = Type
        sheet.cell(row=i, column=8).value = Review4
#ToDo Add Engineering Review and Engineering Review Estimate
Output = 'Changing to Engineering Review Estimate'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(1,sheet.max_row+1):
    EREPlanning = sheet.cell(row=i, column=8).value
    if 'Engineering review estimate'in EREPlanning:
        EREPlanning2 = Planning.split('-')
        Type = 'ENG Rev Est'
        EREPlanning2[0] = ''
        EREPlanning3 = (' ').join(EREPlanning2)
        EREPlanning4 = EREPlanning3.lstrip()
        sheet.cell(row=i, column=7).value = Type
        sheet.cell(row=i, column=8).value = EREPlanning4
#
# Save the second temporary file
#
#print('Saving SNReportTemp2.xlsx.')
Output = 'Saving SNReportTemp2.XLSX'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb.save('SNReportTemp2.xlsx')
#
# Close the file
#
wb.close()
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
# Start Excel with the second temporary file
#
Output = 'Starting Excell with SNReportTemp2.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
os.startfile('SNReportTemp2.xlsx')
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
#
messagebox.showinfo("In Excel", "Sort on Assigned to RT, Engagement RT, Request RT, \n              Task RT, and Estimate RT.\n\nSave as SNReportTemp2.xlsx\n\n\n\nExit Excel!!!!")
#
# Starting the third phase of processing
# Open the second temporary file
#
#print('Opening SNReportTemp2.xls')
Output = 'Opening SNReportTemp2.XLSX'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb = openpyxl.load_workbook('SNReportTemp2.xlsx')
#
# Open the worksheet
#
#print('Opening Worksheet')
Output = 'Opening Worksheet'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
sheet = wb.get_sheet_by_name('Page 1')
#
# Move the estimated hours to the Estimate column on the Planview line
#
#print('Moving Estimated Hours')
Output = 'Moving Estimated Hours'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(2,sheet.max_row+1):
    Planview = sheet.cell(row=i, column=12).value
    if 'Planview work ID'in Planview:
        Estimate2 = sheet.cell(row=i-1, column=10).value
        if Estimate2 != None:
            sheet.cell(row=i, column=12).value = int(Estimate2)
        else:
            sheet.cell(row=i, column=12).value = ""
        for j in range(1,sheet.max_column):
            sheet.cell(row=i-1, column=j).value = ""
#
# Split the PVID cell, write just the PVID back to the cell.
# rebuild the Feature in its own column
#
#print('Moving Feature Information')
Output = 'Moving Feature Information'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for i in range(2,sheet.max_row+1):
    PVIDFeature = sheet.cell(row=i, column=10).value
    if PVIDFeature == None:
        continue
    if PVIDFeature == "":
        continue
    PVF = PVIDFeature.split(' ')
#ToDo Check to see is PVF[0] contains text and handle accordingly
    sheet.cell(row=i, column=10).value = int(PVF[0])
    PVF[0] = ""
    PVFTemp = ' '.join(PVF)
    PVFTemp2 = PVFTemp.lstrip()
    if PVFTemp2 != "":
        # Values for the feature cell
        sheet.cell(row=i, column=11).value = PVFTemp2
    else:
        # No values for the feature cell
        sheet.cell(row=i, column=11).value = 'No Feature'
#
# Save the third Temp file
#
#print('Saving SNReportTemp3.xlsx.')
Output = 'Saving SNReportTemp3.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb.save('SNReportTemp3.xlsx')
#
# Close the file
#
wb.close()
#
# Start Excel with the third temporary file
#
Output = 'Starting Excel with SNReportTemp3.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
os.startfile('SNReportTemp3.xlsx')
#
# Instructions for the tasks to be done in Excel
#
messagebox.showinfo("In Excel", "Sort by Assigned to RT, Engagement RT, \n             Request RT, Task RT to remove empty rows.\n\nSave SNReportTemp3.xlsx\n\n\n\nExit Excel!!!")
#
# Starting the fourth pahse of processing
# Getting data from Finanacial Info and Small Enhancement
#
# Opening the third teporary file
#
#print('Opening SNReportTemp3.xlsx')
Output = 'Opening SNReportTemp3.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb = openpyxl.load_workbook('SNReportTemp3.xlsx')
#Open the worksheet
sheet = wb.get_sheet_by_name('Page 1')
#
# Open Financial Info file
#
#print('opening Financial Info.xlsx')
Output = 'Opening Financial Info.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
fwb = openpyxl.load_workbook('Finance Info.xlsx')
fsheet = fwb.get_sheet_by_name('Sheet1')
#
# Open the Small Enhancement file
#
#print('Opening Small Enhancement')
Output = 'Opening Small Enhancement.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
swb = openpyxl.load_workbook('Small Enhancement.xlsx')
ssheet = swb.get_sheet_by_name('Sheet1')
for i in range(2,sheet.max_row+1):
    PVIDRT = sheet.cell(row=i, column=10).value
    PVIDFeature = sheet.cell(row=i, column=11).value
    if PVIDRT == None:
        sheet.cell(row=i, column=10).value = 'No PVID'
    if PVIDRT == "":
        sheet.cell(row=i, column=11).value = 'No PVID'
    if PVIDFeature == None:
        sheet.cell(row=i, column=11).value = 'No Feature'
    if PVIDFeature == "":
        sheet.cell(row=i, column=11).value = 'No Feature'
#
#  Getting Data from Finacial by PVID and the Small Enhancement File
#
#print('Getting Data from Financial by PVID')
Output = 'Getting Data from Financial Info and Small Enhancement by PVID'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
max_length = 0
for i in range(1,sheet.max_row):
    PVIDRT = str(sheet.cell(row=i, column=10).value)
    if  PVIDRT == 'PVID RT':        # Skip if column header
        continue
    elif PVIDRT == 'None':          #Skip if no PVID
        continue
    elif PVIDRT == '':              #Skip if no PVID
        continue
    elif str(PVIDRT) == '8601':     # PVID is 8601 compare Feature and set the Name
        for s in range(1,ssheet.max_row+1):
            PVIDSE = str(ssheet.cell(row=s, column=1).value)
            if PVIDSE == 'PV #':    #skip if column header
                continue
            if PVIDRT == PVIDSE:    #Look to see if the PVIDRT in in the PVIDST
                PVIDFeature = str(sheet.cell(row=i, column=11).value)        #Read the Feature from SNReportTemp3
                SEFeature = str(ssheet.cell(row=s, column=2).value)          #Read the Feature from Small Enhancement
                if PVIDFeature == SEFeature:                            #Features Match
                    sheet.cell(row=i, column=13).value = ssheet.cell(row=s, column=3).value       #Feature Match, Write value from Small Enhancement Description to Name RT
                    sheet.cell(row=i, column=16).value = ssheet.cell(row=s, column=4).value       #Feature Match, Write value from Small Enhancement Description to Name RT
                    sheet.cell(row=i, column=19).value = ssheet.cell(row=s, column=5).value       #Feature Match, Write value from Small Enhancement Description to Name RT
                    continue
            else:
                continue
    elif str(PVIDRT) == '8576':  # PVID is 8576 compare Feature and set the Name
        for s in range(1, ssheet.max_row + 1):
            PVIDSE = str(ssheet.cell(row=s, column=1).value)
            if PVIDSE == 'PV #':  # skip if column header
                continue
            if PVIDRT == PVIDSE:  # Look to see if the PVIDRT in in the PVIDST
                PVIDFeature = str(sheet.cell(row=i, column=11).value)  # Read the Feature from SNReportTemp3
                SEFeature = str(ssheet.cell(row=s, column=2).value)  # Read the Feature from Small Enhancement
                if PVIDFeature == SEFeature:  # Features Match
                    sheet.cell(row=i, column=13).value = ssheet.cell(row=s, column=3).value  # Feature Match, Write value from Small Enhancement Description to Name RT
                    sheet.cell(row=i, column=16).value = ssheet.cell(row=s, column=4).value       #Feature Match, Write value from Small Enhancement Description to Name RT
                    sheet.cell(row=i, column=19).value = ssheet.cell(row=s, column=5).value       #Feature Match, Write value from Small Enhancement Description to Name RT
                    continue
            else:
                continue
    if PVIDRT !="PVID RT":        # Scan the Finance Info file
        for j in range(1,fsheet.max_row+1):
            PVIDFT = str(fsheet.cell(row=j, column=2).value)
            if PVIDFT == 'PV #':    #skip if column header
                continue
            if PVIDRT in PVIDFT:    #Look to see if the PVIDRT in in the PVIDFT
                sheet.cell(row=i, column=13).value = fsheet.cell(row=j, column=4).value     #Name
                sheet.cell(row=i, column=14).value = fsheet.cell(row=j, column=7).value     #Integration
                sheet.cell(row=i, column=15).value = fsheet.cell(row=j, column=9).value     #Status
                sheet.cell(row=i, column=16).value = fsheet.cell(row=j, column=11).value    #VP
                sheet.cell(row=i, column=17).value = fsheet.cell(row=j, column=12).value    #IT Vertical
                sheet.cell(row=i, column=18).value = fsheet.cell(row=j, column=13).value    #IT Portfolio
                sheet.cell(row=i, column=19).value = fsheet.cell(row=j, column=14).value    #MD
                sheet.cell(row=i, column=20).value = fsheet.cell(row=j, column=15).value    #Sponsor Division
                sheet.cell(row=i, column=21).value = fsheet.cell(row=j, column=16).value    #Sponsor
                sheet.cell(row=i, column=26).value = fsheet.cell(row=j, column=17).value    #Approve Date
                sheet.cell(row=i, column=27).value = fsheet.cell(row=j, column=18).value    #Approve Start
                sheet.cell(row=i, column=28).value = fsheet.cell(row=j, column=19).value    #Approve End
                sheet.cell(row=i, column=29).value = fsheet.cell(row=j, column=20).value    #Planned Start
                sheet.cell(row=i, column=30).value = fsheet.cell(row=j, column=21).value    #Planned End
                sheet.cell(row=i, column=31).value = fsheet.cell(row=j, column=28).value    #Project Manager
                sheet.cell(row=i, column=32).value = fsheet.cell(row=j, column=36).value    #Project Analyst
                sheet.cell(row=i, column=33).value = fsheet.cell(row=j, column=39).value    #Project Category
for i in range(2,sheet.max_row+1):
    FinanceName = sheet.cell(row=i, column=13).value
    if FinanceName == None:
        sheet.cell(row=i, column=13).value = 'xxxxxxxxxx'
        continue
    if FinanceName == "":
        sheet.cell(row=i, column=13).value = 'xxxxxxxxxxz'
#
# Scan all the cells by column and get the maximum cell size for each column
#
#print('Adjusting Column Sizes and Formatting')
Output = 'Adjusting Column Sizes and Formatting'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
for col in sheet.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         if column == 'E':                                      # Format Created RT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'F':                                      # Format Completion RT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'G':                                      #Format Type
             cell.alignment = Alignment(horizontal='center')
         if column == 'H':                                      # Format Description RT
             cell.alignment = Alignment(wrap_text='true')
         if column == 'J':                                      # Format PVID RT
             cell.alignment = Alignment(horizontal='center', wrap_text='true')
         if column == 'K':                                      # Format Feature RT
             cell.alignment = Alignment(horizontal='center', wrap_text='true')
         if column == 'L':                                      # Format Estimate RT
             cell.alignment = Alignment(horizontal='center')
         if column == 'N':                                      # Format Integration FT
             cell.alignment = Alignment(horizontal='center')
         if column == 'O':                                      # Format Status FT
             cell.alignment = Alignment(horizontal='center')
         if column == 'Z':                                      # Format Approve Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AA':                                      # Format Approve Start Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AB':                                      # Format Approve End Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AC':                                     # Format Planned Start Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AD':                                     # Format Planned End Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         try:                                                   # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2                    # Define the column width
     sheet.column_dimensions[column].width = adjusted_width     # Set the column width
#
# set columns to a specified width instead of maximum
#
#print('Adjusting specific column sizes')
Output = 'Adjusting specific column sizes'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
sheet.column_dimensions['B'].width = 14                 # Engagement RT
sheet.column_dimensions['C'].width = 14                 # Request RT
sheet.column_dimensions['D'].width = 15                 # Task RT
sheet.column_dimensions['E'].width = 14                 # Created RT
sheet.column_dimensions['F'].width = 14                 # Completion RT
sheet.column_dimensions['G'].width = 10                 # Type RT
sheet.column_dimensions['H'].width = 76                 # Description RT
sheet.column_dimensions['I'].width = 18                 # Status RT
sheet.column_dimensions['J'].width = 10                 # PVID RT
sheet.column_dimensions['K'].width = 20                 # Feature RT
sheet.column_dimensions['L'].width = 11                 # Estimate RT
sheet.column_dimensions['O'].width = 40                 # Status FT
#
# Saving the fourth temporary file
#
Output = 'Saving SNReportFTemp.xlsx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
wb.save('SNReportFTemp.xlsx')
# Close the file
wb.close()
#
# Opening fourth temporary file in Excel with instructions
#
Output = 'Starting Excel with SNReportFTemp.xslx'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
os.startfile('SNReportFTemp.xlsx')
messagebox.showinfo("In Excel", "Save to Projected Time Folder with appropriate name.\n\nExit Excel")
#
# Delete the temporary files
#
Output = 'Removing temporary files'
LABEL = Label(root, text=Output)
LABEL.pack()
root.update()
os.remove('SNReportTemp.xlsx')
os.remove('SNReportTemp2.xlsx')
os.remove('SNReportTemp3.xlsx')
os.remove('SNReportFTemp.xlsx')
messagebox.showinfo("In Python", "Service Now Report processing Complete!")
