#! python3
#
# SNReport3L.py -- Read the Service Now Assessment Report
# Three Line Report version
#
# import OPENPYXL Functions
#
import openpyxl
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.comments import Comment
#
# Import OS Functions
#
import os
#
# Import TK GUI FUNCTIONS
#
import tkinter
from tkinter import messagebox
from tkinter import filedialog
#
# Starting TK and hiding the main window
#
root = tkinter.Tk()
root.withdraw()
#
# Starting the first phase of processing
# Get the input file name to process
#
SNFile = filedialog.askopenfile(parent=root,mode='rb',title='Enter the File Name to process')
print('Opening Workbook', SNFile.name)
#
# Open the input file
#
wb = openpyxl.load_workbook(SNFile)
#
# Open the worksheet
#
print('Opening Worksheet')
sheet = wb.get_sheet_by_name('Page 1')
#
# Save the temporary file
#
print('Saving SNReportTemp.xlsx.')
wb.save('SNReportTemp.xlsx')
#
# Close the temporary file
#
wb.close()
#
# Start Excel with the temporary file
# Instructions for the expected Excel error
#
messagebox.showinfo("Starting Excel", "Answer YES to the Error, and then close the Error Log Box")
os.startfile('SNReportTemp.xlsx')
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
#
messagebox.showinfo("In Excel", "Remove columns B, D, F, H.")
messagebox.showinfo("In Excel", "Cut Column E and Insert at D.")
messagebox.showinfo("In Excel", "Insert One column at I.")
messagebox.showinfo("In Excel", "Starting at column K, Insert 9 Columns (K thru S).")
messagebox.showinfo("In Excel", "Save as SNReportTemp.xlsx, Overwriting the file!")
messagebox.showwarning("In Excel", "Exit Excel")
#
# Start the second phase of processing
# Open the temporary file
#
print('Opening File SNReportTemp.xlsx')
wb = openpyxl.load_workbook('SNReportTemp.xlsx')
#
# Open the worksheet
#
print('Opening Worksheet')
sheet = wb.get_sheet_by_name('Page 1')
#
# Adding the names for the title row
#
print('Adding New Row 1 Names')
sheet['a1'] = 'Assigned To RT'
sheet['b1'] = 'Engagement RT'
sheet['c1'] = 'Request RT'
sheet['d1'] = 'Task RT'
sheet['e1'] = 'Created RT'
sheet['f1'] = 'Description RT'
sheet['g1'] = 'Status RT'
sheet['h1'] = 'PVID RT'
sheet['i1'] = 'Feature RT'
sheet['j1'] = 'Estimate RT'
sheet['k1'] = 'Name FT'
sheet['l1'] = 'Integration FT'
sheet['m1'] = 'Status FT'
sheet['n1'] = 'VP FT'
sheet['o1'] = 'IT Vertical FT'
sheet['p1'] = 'IT Portfolio FT'
sheet['q1'] = 'MD FT'
sheet['r1'] = 'Sponsor Division FT'
sheet['s1'] = 'Sponsor FT'
sheet['t1'] = 'Requested By Name RT'
sheet['u1'] = 'Requested By Email RT'
sheet['v1'] = 'Requested For Name RT'
sheet['w1'] = 'Requested For Email RT'
sheet['x1'] = 'Approve Date FT'
sheet['y1'] = 'Approve Start Date FT'
sheet['z1'] = 'Approve End Date FT'
sheet['aa1'] = 'Planned Start Date FT'
sheet['ab1'] = 'Planned End Date FT'
sheet['ac1'] = 'Project Manager FT'
sheet['ad1'] = 'Project Analyst FT'
sheet['ae1'] = 'Project Category FT'
sheet = wb.get_sheet_by_name('Page 1')
#
# Scan all the cells by column and get the maximum cell size for each column
#
print('Adjusting Column Sizes and Formatting')
for col in sheet.columns:
     max_length = 0
     column = col[0].column # Get the column name
     if column == 'E':                                  # Format Created RT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy h:mm AM/PM'
     if column == 'F':                                  # Format Description RT
         cell.alignment = Alignment(wrap_text='true')
     if column == 'H':                                  # Format PVID RT
         cell.alignment = Alignment(horizontal='center', wrap_text='true')
     if column == 'I':                                  # Format Feature RT
         cell.alignment = Alignment(horizontal='center', wrap_text='true')
     if column == 'J':                                  # Format Estimate RT
         cell.alignment = Alignment(horizontal='center')
     if column == 'X':                                  # Format Approve Date FT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy'
     if column == 'Y':                                  # Format Approve Start Date FT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy'
     if column == 'Z':                                  # Format Approve End Date FT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy'
     if column == 'AA':                                 # Format Planned Start Date FT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy'
     if column == 'AB':                                 # Format Planned End Date FT
         cell.alignment = Alignment(horizontal='center')
         cell.number_format = 'mm/dd/yy'
     for cell in col:
         try:                                           # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2                    # Define the column width
     sheet.column_dimensions[column].width = adjusted_width     # Set the column width
#
# set the fixed column widths
#
print('Adjusting specific column sizes')
sheet.column_dimensions['E'].width = 18                 # Created RT
sheet.column_dimensions['F'].width = 76                 # Description RT
sheet.column_dimensions['H'].width = 24                 # PVID RT
sheet.column_dimensions['G'].width = 18                 # Status RT
sheet.column_dimensions['I'].width = 20                 # Feature RT
sheet.column_dimensions['J'].width = 16                 # Estimate RT
#
# Changing the beginning of the Descriotion to the short version
#
# Change Deliver
print('Changing to Deliver')
for i in range(1,sheet.max_row+1):
    Description = sheet.cell(row=i, column=6).value
    if 'Deliver network assessment'in Description:
        Description2 = Description.split('-')
        Description2[0] = 'Deliver '
        Description3 = ('-').join(Description2)
        sheet.cell(row=i, column=6).value = Description3
# Change Estimate
print('Changing to Estimate')
for i in range(1,sheet.max_row+1):
    Description = sheet.cell(row=i, column=6).value
    if 'Build assessment estimate'in Description:
        Description2 = Description.split('-')
        Description2[0] = 'Estimate '
        Description3 = ('-').join(Description2)
        sheet.cell(row=i, column=6).value = Description3
# Change Planning
print('Changing to Planning')
for i in range(1,sheet.max_row+1):
    Description = sheet.cell(row=i, column=6).value
    if 'Resource & Demand planning for'in Description:
        Description2 = Description.split('-')
        Description2[0] = 'Planning '
        Description3 = ('-').join(Description2)
        sheet.cell(row=i, column=6).value = Description3
#
# Save the second temporary file
#
print('Saving SNReportTemp2.xlsx.')
wb.save('SNReportTemp2.xlsx')
#
# Close the file
#
wb.close()
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
# Start Excel with the second temporary file
#
os.startfile('SNReportTemp2.xlsx')
#
# Instructions for the tasks to be done in Excel
# Instruction message box display
#
messagebox.showinfo("In Excel", "Sort the data on Assigned to RT, Engagement RT, Request RT, Task RT, and Estimate RT.")
messagebox.showinfo("In Excel", "Resize the Rows after the sort.")
messagebox.showinfo("In Excel", "Save as SNReportTemp2.xlsx")
messagebox.showwarning("In Excel", "Exit Excel")
#
# Starting the third phase of processing
# Open the second temporary file
#
print('Opening SNReportTemp2.xls')
wb = openpyxl.load_workbook('SNReportTemp2.xlsx')
#
# Open the worksheet
#
print('Opening Worksheet')
sheet = wb.get_sheet_by_name('Page 1')
#
# Move the estimated hours to the Estimate column on the Planview line
#
print('Moving Estimated Hours')
for i in range(2,sheet.max_row+1):
    Planview = sheet.cell(row=i, column=10).value
    if 'Planview work ID'in Planview:
        Estimate2 = sheet.cell(row=i-1, column=8).value
        if Estimate2 != None:
            sheet.cell(row=i, column=10).value = int(Estimate2)
        else:
            sheet.cell(row=i, column=10).value = ""
        for j in range(1,sheet.max_column):
            sheet.cell(row=i-1, column=j).value = ""
            sheet.cell(row=i+1, column=j).value = ""
#
# Split the PVID cell, write just the PVID back to the cell.
# rebuild the Feature in its own column
#
print('Moving Feature Information')
for i in range(2,sheet.max_row+1):
    PVIDFeature = sheet.cell(row=i, column=8).value
    if PVIDFeature == None:
        continue
    if PVIDFeature == "":
        continue
    PVF = PVIDFeature.split(' ')
    sheet.cell(row=i, column=8).value = int(PVF[0])
    PVF[0] = ""
    PVFTemp = ' '.join(PVF)
    PVFTemp2 = PVFTemp.lstrip()
    sheet.cell(row=i, column=9).value = PVFTemp2
#
# Save the third Temp file
#
print('Saving SNReportTemp3.xlsx.')
wb.save('SNReportTemp3.xlsx')
#
# Close the file
#
wb.close()
#
# Start Excel with the third temporary file
#
os.startfile('SNReportTemp3.xlsx')
#
# Instructions for the tasks to be done in Excel
#
messagebox.showinfo("In Excel", "Sort by Assigned to RT, Engagement RT, Request RT, Task RT to remove empty rows")
messagebox.showinfo("In Excel", "Resize the Rows after the sort.")
messagebox.showinfo("In Excel", "Save SNReportTemp3.xlsx")
messagebox.showwarning("In Excel", "Exit Excel")
#
# Starting the fourth pahse of processing
# Getting data from Finanacial Info and Small Enhancement
#
# Opening the third teporary file
#
print('Opening SNReportTemp3.xlsx')
wb = openpyxl.load_workbook('SNReportTemp3.xlsx')
#Open the worksheet
sheet = wb.get_sheet_by_name('Page 1')
#
# Open Financial Info file
#
print('opening Financial Info.xlsx')
fwb = openpyxl.load_workbook('Finance Info.xlsx')
fsheet = fwb.get_sheet_by_name('Sheet1')
#
# Open the Small Enhancement file
#
print('Opening Small Enhancement')
swb = openpyxl.load_workbook('Small Enhancement.xlsx')
ssheet = fwb.get_sheet_by_name('Sheet1')
#
#  Getting Data from Finacial by PVID and the Small Enhancement File
#
print('Getting Data from Financial by PVID')
max_length = 0
for i in range(1,sheet.max_row):
    PVIDRT = str(sheet.cell(row=i, column=8).value)
    if  PVIDRT == 'PVID RT':        # Skip if column header
        continue
    elif PVIDRT == 'None':          #Skip if no PVID
        continue
    elif str(PVIDRT) == '8601':     # PVID is 8601 compare Feature and set the Name
        PVIDFeature = sheet.cell(row=i, column=9).value
        if PVIDFeature != 'None':
            if PVIDFeature == 'F3607':
                sheet.cell(row=i, column=11).value = 'NAIA'
            if PVIDFeature == 'F20206':
                sheet.cell(row=i, column=11).value = 'NAIA - CRE'
            if PVIDFeature == 'F20207':
                sheet.cell(row=i, column=11).value = 'NAIA - Network Engineering'
            if PVIDFeature == 'F20205':
                sheet.cell(row=i, column=11).value = 'NAIA - Operational Nexessity'
            if PVIDFeature == 'None':
                sheet.cell(row=i, column=11).value = 'Feature Not Defined'
            else:
                continue
    elif str(PVIDRT) == '8576':     # PVID is 8576 compare Feature and set the Name
        PVIDFeature = sheet.cell(row=i, column=9).value
        if PVIDFeature != 'None':
            if PVIDFeature == 'F20205':
                sheet.cell(row=i, column=11).value = 'NAIA'
            if PVIDFeature == 'None':
                sheet.cell(row=i, column=11).value = 'Feature Not Defined'
            if PVIDFeature == 'F':
                sheet.cell(row=i, column=11).value = 'Feature Not Defined'
            else:
                continue
    if PVIDRT !="PVID RT":        # Scan the Finance Info file
        for j in range(1,fsheet.max_row+1):
            PVIDFT = str(fsheet.cell(row=j, column=2).value)
            if PVIDFT == 'PV #':    #skip if column header
                continue
            if PVIDRT in PVIDFT:    #Look to see if the PVIDRT in in the PVIDFT
                sheet.cell(row=i, column=11).value = fsheet.cell(row=j, column=4).value     #Name
                sheet.cell(row=i, column=12).value = fsheet.cell(row=j, column=7).value     #Integration
                sheet.cell(row=i, column=13).value = fsheet.cell(row=j, column=9).value     #Status
                sheet.cell(row=i, column=14).value = fsheet.cell(row=j, column=11).value    #VP
                sheet.cell(row=i, column=15).value = fsheet.cell(row=j, column=12).value    #IT Vertical
                sheet.cell(row=i, column=16).value = fsheet.cell(row=j, column=13).value    #IT Portfolio
                sheet.cell(row=i, column=17).value = fsheet.cell(row=j, column=14).value    #MD
                sheet.cell(row=i, column=18).value = fsheet.cell(row=j, column=15).value    #Sponsor Division
                sheet.cell(row=i, column=19).value = fsheet.cell(row=j, column=16).value    #Sponsor
                sheet.cell(row=i, column=24).value = fsheet.cell(row=j, column=17).value    #Approve Date
                sheet.cell(row=i, column=25).value = fsheet.cell(row=j, column=18).value    #Approve Start
                sheet.cell(row=i, column=26).value = fsheet.cell(row=j, column=19).value    #Approve End
                sheet.cell(row=i, column=27).value = fsheet.cell(row=j, column=20).value    #Planned Start
                sheet.cell(row=i, column=28).value = fsheet.cell(row=j, column=21).value    #Planned End
                sheet.cell(row=i, column=29).value = fsheet.cell(row=j, column=28).value    #Project Manager
                sheet.cell(row=i, column=31).value = fsheet.cell(row=j, column=39).value    #Project Category
#
# Scan all the cells by column and get the maximum cell size for each column
#
print('Adjusting Column Sizes and Formatting')
for col in sheet.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         if column == 'E':                                      # Format Created RT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy h:mm AM/PM'
         if column == 'F':                                      # Format Description RT
             cell.alignment = Alignment(wrap_text='true')
         if column == 'H':                                      # Format PVID RT
             cell.alignment = Alignment(horizontal='center', wrap_text='true')
         if column == 'I':                                      # Format Feature RT
             cell.alignment = Alignment(horizontal='center', wrap_text='true')
         if column == 'J':                                      # Format Estimate RT
             cell.alignment = Alignment(horizontal='center')
         if column == 'X':                                      # Format Approve Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'Y':                                      # Format Approve Start Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'Z':                                      # Format Approve End Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AA':                                     # Format Planned Start Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         if column == 'AB':                                     # Format Planned End Date FT
             cell.alignment = Alignment(horizontal='center')
             cell.number_format = 'mm/dd/yy'
         try:                                                   # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2                    # Define the column width
     sheet.column_dimensions[column].width = adjusted_width     # Set the column width
#
# set columns to a specified width instead of maximum
#
print('Adjusting specific column sizes')
sheet.column_dimensions['E'].width = 18                 # Created RT
sheet.column_dimensions['F'].width = 76                 # Description RT
sheet.column_dimensions['G'].width = 18                 # Status RT
sheet.column_dimensions['H'].width = 24                 # PVID RT
sheet.column_dimensions['I'].width = 20                 # Feature RT
sheet.column_dimensions['J'].width = 16                 # Estimate RT
#
# Saving the fourth temporary file
#
wb.save('SNReportFTemp.xlsx')
# Close the file
wb.close()
#
# Opening fourth temporary file in Excel with instructions
#
os.startfile('SNReportFTemp.xlsx')
messagebox.showinfo("In Excel", "Save to Projected Time Folder with appropriate name")
messagebox.showwarning("In Excel", "Exit Excel")
messagebox.showinfo("In Python", "Service Now Report processing Complete!")
#
# Delete the temporary files
#
os.remove('SNReportTemp.xlsx')
os.remove('SNReportTemp2.xlsx')
os.remove('SNReportTemp3.xlsx')
os.remove('SNReportFTemp.xlsx')
